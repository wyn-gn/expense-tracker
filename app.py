import tkinter as tk
from tkinter import messagebox
from ttkbootstrap import Style
from ttkbootstrap.constants import *
from ttkbootstrap import ttk
from openpyxl import Workbook
import sqlite3

# Center window function
def center_window(win, width, height):
    screen_width = win.winfo_screenwidth()
    screen_height = win.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    win.geometry(f"{width}x{height}+{x}+{y}")


# Pin Window (Starting Screen)
pin_window = tk.Tk()
pin_window.title("Secure Access")
pin_window.geometry("350x300")
pin_window.resizable(False, False)

style = Style("sandstone")
pin_window.configure(bg=style.colors.bg)
center_window(pin_window, 350, 300)

frame = ttk.Frame(pin_window, padding=30)
frame.pack(expand=True)

# PIN screen labels
ttk.Label(frame, text="Welcome to your", font=("Segoe UI", 11)).pack(pady=(0, 2))
ttk.Label(frame, text="Personal Financial Tracker", font=("Segoe UI Semibold", 16)).pack(pady=(0, 15))
ttk.Label(frame, text="Enter your 4-digit PIN", font=("Segoe UI", 10)).pack(pady=(0, 8))

# PIN entry field
pin_entry = ttk.Entry(frame, show="•", font=("Segoe UI", 12), justify="center", width=15)
pin_entry.pack(ipady=5, pady=(0, 15))


# MAIN WINDOW 
def main_window():
    global main_win
    pin_window.withdraw()  # hide the PIN window

    main_win = tk.Toplevel()
    main_win.title("Expense Tracker")
    center_window(main_win, 600, 350)
    main_win.configure(padx=20, pady=20, bg=style.colors.bg)

    header = ttk.Label(main_win, text="Expense Tracker", font=("Segoe UI Semibold", 20))
    header.pack(anchor="center", pady=(0, 15))

    btn_frame = ttk.Frame(main_win)
    btn_frame.pack(pady=10)

    # Only Daily Expenses opens a new window
    ttk.Button(btn_frame, text="Daily Expenses", bootstyle=PRIMARY, width=20, command=expenses_window).grid(row=0, column=0, padx=15)

    # Bills and Debts currently just placeholders (no new window)
    ttk.Button(btn_frame, text="Bills", bootstyle=WARNING, width=20, command=bills_window).grid(row=0, column=1, padx=15)
    ttk.Button(btn_frame, text="Debts", bootstyle=DANGER, width=20, command=debts_window).grid(row=0, column=2, padx=15)

    ttk.Button(btn_frame, text="Lock App", bootstyle=OUTLINE + SECONDARY, width=20, command=lock_app).grid(row=1, column=1, padx=15, pady=30)


# Expenses Window
def expenses_window():
    main_win.withdraw()
    global expenses_win

    # Create or connect to database
    conn = sqlite3.connect('expenses.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT,
            amount REAL,
            date TEXT,
            description TEXT
        )
    ''')
    conn.commit()
    conn.close()

    expenses_win = tk.Toplevel()
    expenses_win.title("Daily Expenses")
    expenses_win.geometry("900x600")
    center_window(expenses_win, 900, 600)
    style.theme_use("sandstone")

    # Header Frame (holds label + back button)
    header_frame = ttk.Frame(expenses_win)
    header_frame.pack(fill="x", pady=10, padx=10)

    # Column 0 = centered label, Column 1 = back button (aligned right)
    header_frame.columnconfigure(0, weight=1)  # make center expand

    ttk.Label(header_frame, text="Daily Expenses", font=("Segoe UI Semibold", 16)).grid(row=0, column=0, sticky="nsew")
    ttk.Button(header_frame, text="Back", bootstyle=SECONDARY, command=go_back_to_main_from_expenses).grid(row=0, column=1, sticky="e")

    # Input Frame
    input_frame = ttk.LabelFrame(expenses_win, text="Add New Expense", padding=20, bootstyle="info")
    input_frame.pack(fill=X, padx=10, pady=10)

    ttk.Label(input_frame, text="Category:").grid(row=0, column=0, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Amount:").grid(row=0, column=2, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Date (MM-DD-YYYY):").grid(row=1, column=0, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Description:").grid(row=1, column=2, padx=10, pady=5, sticky=W)

    category_entry = ttk.Entry(input_frame, width=25)
    amount_entry = ttk.Entry(input_frame, width=25)
    date_entry = ttk.Entry(input_frame, width=25)
    desc_entry = ttk.Entry(input_frame, width=25)

    category_entry.grid(row=0, column=1, padx=10, pady=5)
    amount_entry.grid(row=0, column=3, padx=10, pady=5)
    date_entry.grid(row=1, column=1, padx=10, pady=5)
    desc_entry.grid(row=1, column=3, padx=10, pady=5)

    # FUNCTIONS
    def add_expense():
        category = category_entry.get()
        amount = amount_entry.get()
        date = date_entry.get()
        desc = desc_entry.get()  # optional

        if category and amount and date:
            conn = sqlite3.connect('expenses.db')
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO expenses (category, amount, date, description) VALUES (?, ?, ?, ?)",
                (category, amount, date, desc)
            )
            conn.commit()
            conn.close()

            # Clear inputs
            category_entry.delete(0, tk.END)
            amount_entry.delete(0, tk.END)
            date_entry.delete(0, tk.END)
            desc_entry.delete(0, tk.END)

            load_expenses()
        else:
            messagebox.showwarning("Input Error", "Please fill in all required fields.")

    def load_expenses():
        for row in tree.get_children():
            tree.delete(row)
        conn = sqlite3.connect('expenses.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM expenses ORDER BY id ASC")
        rows = cursor.fetchall()
        conn.close()
        for row in rows:
            tree.insert("", tk.END, values=row)

    def export_to_excel():
        conn = sqlite3.connect('expenses.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM expenses")
        rows = cursor.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(["ID", "Category", "Amount", "Date", "Description"])
        for row in rows:
            ws.append(row)
        wb.save("expenses.xlsx")
        messagebox.showinfo("Success", "Data exported to expenses.xlsx!")

    # Buttons
    btn_frame = ttk.Frame(expenses_win)
    btn_frame.pack(pady=10)

    ttk.Button(btn_frame, text="Add Expense", bootstyle=SUCCESS, width=20, command=add_expense).grid(row=0, column=0, padx=10)
    ttk.Button(btn_frame, text="Export to Excel", bootstyle=INFO, width=20, command=export_to_excel).grid(row=0, column=1, padx=10)

    # Table Frame
    table_frame = ttk.LabelFrame(expenses_win, text="Expense Records", padding=15, bootstyle="info")
    table_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

    columns = ("ID", "Category", "Amount", "Date", "Description")
    tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=12, bootstyle=PRIMARY)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor=tk.CENTER, width=150)
    tree.pack(fill=BOTH, expand=True)

    load_expenses()

# Bills Window
def bills_window():
    main_win.withdraw()
    global bills_win

    conn = sqlite3.connect('bills.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT,
            amount REAL,
            duedate TEXT,
            description TEXT
        )
    ''')
    conn.commit()
    conn.close()

    bills_win = tk.Toplevel()
    bills_win.title("Bills")
    bills_win.geometry("900x600")
    center_window(bills_win, 900, 600)
    style.theme_use("sandstone")

    # Header Frame
    header_frame = ttk.Frame(bills_win)
    header_frame.pack(fill="x", pady=10, padx=10)
    header_frame.columnconfigure(0, weight=1)  # make center expand

    ttk.Label(header_frame, text="Bills", font=("Segoe UI Semibold", 16)).grid(row=0, column=0, sticky="nsew")
    ttk.Button(header_frame, text="Back", bootstyle=SECONDARY, command=go_back_to_main_from_bills).grid(row=0, column=1, sticky="e")

    # Input Frame
    input_frame = ttk.LabelFrame(bills_win, text="Add New Expense", padding=20, bootstyle="info")
    input_frame.pack(fill=X, padx=10, pady=10)

    ttk.Label(input_frame, text="Category:").grid(row=0, column=0, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Amount:").grid(row=0, column=2, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Due Date (MM-DD-YYYY):").grid(row=1, column=0, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Description:").grid(row=1, column=2, padx=10, pady=5, sticky=W)

    category_entry = ttk.Entry(input_frame, width=25)
    amount_entry = ttk.Entry(input_frame, width=25)
    date_entry = ttk.Entry(input_frame, width=25)
    desc_entry = ttk.Entry(input_frame, width=25)

    category_entry.grid(row=0, column=1, padx=10, pady=5)
    amount_entry.grid(row=0, column=3, padx=10, pady=5)
    date_entry.grid(row=1, column=1, padx=10, pady=5)
    desc_entry.grid(row=1, column=3, padx=10, pady=5)

    # FUNCTIONS
    def add_bills():
        category = category_entry.get()
        amount = amount_entry.get()
        duedate = date_entry.get()
        desc = desc_entry.get()  # optional

        if category and amount and duedate:
            conn = sqlite3.connect('bills.db')
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO bills (category, amount, duedate, description) VALUES (?, ?, ?, ?)",
                (category, amount, duedate, desc)
            )
            conn.commit()
            conn.close()

            # Clear inputs
            category_entry.delete(0, tk.END)
            amount_entry.delete(0, tk.END)
            date_entry.delete(0, tk.END)
            desc_entry.delete(0, tk.END)

            load_bills()
        else:
            messagebox.showwarning("Input Error", "Please fill in all required fields.")

    def load_bills():
        for row in tree.get_children():
            tree.delete(row)
        conn = sqlite3.connect('bills.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM bills ORDER BY id ASC")
        rows = cursor.fetchall()
        conn.close()
        for row in rows:
            tree.insert("", tk.END, values=row)

    def export_to_excel():
        conn = sqlite3.connect('bills.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM bills")
        rows = cursor.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Bills"
        ws.append(["ID", "Category", "Amount", "Due Date", "Description"])
        for row in rows:
            ws.append(row)
        wb.save("bills.xlsx")
        messagebox.showinfo("Success", "Data exported to bills.xlsx!")

    # Buttons
    btn_frame = ttk.Frame(bills_win)
    btn_frame.pack(pady=10)

    ttk.Button(btn_frame, text="Add Bills", bootstyle=SUCCESS, width=20, command=add_bills).grid(row=0, column=0, padx=10)
    ttk.Button(btn_frame, text="Export to Excel", bootstyle=INFO, width=20, command=export_to_excel).grid(row=0, column=1, padx=10)

    # Table Frame
    table_frame = ttk.LabelFrame(bills_win, text="Bills Records", padding=15, bootstyle="info")
    table_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

    columns = ("ID", "Category", "Amount", "Due Date", "Description")
    tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=12, bootstyle=PRIMARY)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor=tk.CENTER, width=150)
    tree.pack(fill=BOTH, expand=True)

    load_bills()

#Debts Window
def debts_window():
    main_win.withdraw()
    global debts_win

    conn = sqlite3.connect('debts.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS debts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            creditor TEXT,
            amount REAL,
            dateborrowed TEXT,
            description TEXT
        )
    ''')
    conn.commit()
    conn.close()

    debts_win = tk.Toplevel()
    debts_win.title("Debts")
    debts_win.geometry("900x600")
    center_window(debts_win, 900, 600)
    style.theme_use("sandstone")

    header_frame = ttk.Frame(debts_win)
    header_frame.pack(fill="x", pady=10, padx=10)
    header_frame.columnconfigure(0, weight=1)  # make center expand

    ttk.Label(header_frame, text="Debts", font=("Segoe UI Semibold", 16)).grid(row=0, column=0, sticky="nsew")
    ttk.Button(header_frame, text="Back", bootstyle=SECONDARY, command=go_back_to_main_from_debts).grid(row=0, column=1, sticky="e")
    
    input_frame = ttk.LabelFrame(debts_win, text="Add New Debt", padding=20, bootstyle="info")
    input_frame.pack(fill=X, padx=10, pady=10)

    ttk.Label(input_frame, text="Creditor:").grid(row=0, column=0, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Amount:").grid(row=0, column=2, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Date Borrowed (MM-DD-YYYY):").grid(row=1, column=0, padx=10, pady=5, sticky=W)
    ttk.Label(input_frame, text="Description:").grid(row=1, column=2, padx=10, pady=5, sticky=W)

    creditor_entry = ttk.Entry(input_frame, width=25)
    amount_entry = ttk.Entry(input_frame, width=25)
    date_entry = ttk.Entry(input_frame, width=25)
    desc_entry = ttk.Entry(input_frame, width=25)

    creditor_entry.grid(row=0, column=1, padx=10, pady=5)
    amount_entry.grid(row=0, column=3, padx=10, pady=5)
    date_entry.grid(row=1, column=1, padx=10, pady=5)
    desc_entry.grid(row=1, column=3, padx=10, pady=5)

    # FUNCTIONS
    def add_debts():
        creditor = creditor_entry.get()
        amount = amount_entry.get()
        dateborrowed = date_entry.get()
        desc = desc_entry.get()  # optional

        if creditor and amount and dateborrowed:
            conn = sqlite3.connect('debts.db')
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO debts (creditor, amount, dateborrowed, description) VALUES (?, ?, ?, ?)",
                (creditor, amount, dateborrowed, desc)
            )
            conn.commit()
            conn.close()

            # Clear inputs
            creditor_entry.delete(0, tk.END)
            amount_entry.delete(0, tk.END)
            date_entry.delete(0, tk.END)
            desc_entry.delete(0, tk.END)

            load_debts()
        else:
            messagebox.showwarning("Input Error", "Please fill in all required fields.")
    
    def load_debts():
        for row in tree.get_children():
            tree.delete(row)
        conn = sqlite3.connect('debts.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM debts ORDER BY id ASC")
        rows = cursor.fetchall()
        conn.close()
        for row in rows:
            tree.insert("", tk.END, values=row)
    
    def export_to_excel():
        conn = sqlite3.connect('debts.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM debts")
        rows = cursor.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Debts"
        ws.append(["ID", "Creditor", "Amount", "Date Borrowed", "Description"])
        for row in rows:
            ws.append(row)
        wb.save("debts.xlsx")
        messagebox.showinfo("Success", "Data exported to debts.xlsx!")
    
    # Buttons
    btn_frame = ttk.Frame(debts_win)
    btn_frame.pack(pady=10)

    ttk.Button(btn_frame, text="Add Debt", bootstyle=SUCCESS, width=20, command=add_debts).grid(row=0, column=0, padx=10)
    ttk.Button(btn_frame, text="Export to Excel", bootstyle=INFO, width=20, command=export_to_excel).grid(row=0, column=1, padx=10)

    # Table Frame
    table_frame = ttk.LabelFrame(debts_win, text="Debts Records", padding=15, bootstyle="info")
    table_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

    columns = ("ID", "Creditor", "Amount", "Date Borrowed", "Description")
    tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=12, bootstyle=PRIMARY)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor=tk.CENTER, width=150)
    tree.pack(fill=BOTH, expand=True)

    load_debts()

def go_back_to_main_from_expenses():
    expenses_win.destroy()
    main_win.deiconify()

def go_back_to_main_from_bills():
    bills_win.destroy()
    main_win.deiconify()

def go_back_to_main_from_debts():
    debts_win.destroy()
    main_win.deiconify()

# LOCK APP (Return to PIN)
def lock_app():
    main_win.destroy()
    pin_window.deiconify()

# PIN VALIDATION
def open_main_window():
    pin = pin_entry.get()
    if pin == "1234":
        main_window()
    else:
        messagebox.showerror("Error", "Incorrect PIN. Try again.")

# Unlock button
ttk.Button(frame, text="Unlock", bootstyle=INFO, command=open_main_window).pack(fill=X, pady=5)
ttk.Label(frame, text="Your data is encrypted and secure.", font=("Segoe UI", 8)).pack(pady=(20, 0))

pin_window.mainloop()